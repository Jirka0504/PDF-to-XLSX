from __future__ import annotations
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging
from .types import LineItem

log = logging.getLogger(__name__)

DEFAULT_MAPPING = {
    # column name in template : attribute on LineItem
    "Číslo produktu": "product_number",
    "Označení produktu": "product_name",
    "Celní kód zboží": "customs_code",
    "Hmotnost v gramech": "weight_g",
    "Dodané množství": "delivered_qty",
    "Čistá cena za jednotku": "net_unit_price",
    "Celková cena za množství produktu": "total_price",
}

def _find_header_row(ws: Worksheet, required_headers: List[str], scan_rows: int = 50) -> int:
    for r in range(1, scan_rows + 1):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if all(h in row_vals for h in required_headers):
            return r
    return -1

def write_items_to_template(template_path: str, output_path: str, items: List[LineItem], options: Dict[str, Any]) -> None:
    wb = load_workbook(template_path)
    sheet_name = options.get("sheet_name") or wb.sheetnames[0]
    ws = wb[sheet_name]

    mapping: Dict[str, str] = options.get("mapping") or DEFAULT_MAPPING
    headers = list(mapping.keys())

    header_row = _find_header_row(ws, headers)
    if header_row == -1:
        raise ValueError(f"Cannot find header row with required headers in sheet '{sheet_name}'. Required: {headers}")

    # Map header -> column index
    header_cells = ws[header_row]
    header_to_col = {}
    for idx, cell in enumerate(header_cells, start=1):
        val = str(cell.value).strip() if cell.value is not None else ""
        if val in mapping:
            header_to_col[val] = idx

    start_row = header_row + 1

    # Optional: clear existing data (until first empty line in first mapped column)
    if options.get("clear_existing", True):
        max_rows = ws.max_row
        first_col = header_to_col[headers[0]]
        for r in range(start_row, max_rows + 1):
            if ws.cell(r, first_col).value in (None, ""):
                break
            for h in headers:
                ws.cell(r, header_to_col[h]).value = None

    # Write items
    r = start_row
    for it in items:
        for h, attr in mapping.items():
            c = header_to_col[h]
            ws.cell(r, c).value = getattr(it, attr, "")
        r += 1

    wb.save(output_path)
    log.info("Saved output: %s", output_path)
